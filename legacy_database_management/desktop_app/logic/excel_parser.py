"""Excel Parser Module - Scans directories and parses Excel/CSV files"""

import os
import sys
import csv
import codecs
from pathlib import Path
from typing import List, Dict, Any, Optional, Iterator
import openpyxl
import xlrd


# ============ Excel Parser Core ============

def stat_mtime_ms(stat_result: os.stat_result) -> int:
    """Return filesystem modified time as integer milliseconds."""
    return int(stat_result.st_mtime_ns // 1_000_000)

def is_excel_file(filepath: str) -> bool:
    """Check if file has an Excel or CSV extension"""
    excel_exts = {'.xlsx', '.xlsm', '.xlsb', '.xls', '.csv'}
    return Path(filepath).suffix.lower() in excel_exts


def scan_directory(root_path: str, recursive: bool = True) -> List[str]:
    """Recursively scan directory for Excel/CSV files"""
    results = []
    root = Path(root_path)

    if not root.exists():
        return results

    try:
        if recursive:
            for file_path in root.rglob('*'):
                if is_excel_file(str(file_path)):
                    results.append(str(file_path))
        else:
            for file_path in root.iterdir():
                if file_path.is_file() and is_excel_file(str(file_path)):
                    results.append(str(file_path))
    except Exception as e:
        print(f"Error scanning directory {root_path}: {e}", file=sys.stderr)

    return sorted(results)


def read_excel_xlsx(filepath: str) -> Dict[str, List[List[Any]]]:
    """Read .xlsx/.xlsm/.xlsb using openpyxl"""
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True, keep_links=False)
        sheets_data = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                # Keep empty rows as None but we'll filter later
                rows.append([cell for cell in row])
            sheets_data[sheet_name] = rows
        wb.close()
        return sheets_data
    except Exception as e:
        raise RuntimeError(f"Failed to read {filepath}: {e}")


def read_excel_xls(filepath: str) -> Dict[str, List[List[Any]]]:
    """Read legacy .xls using xlrd"""
    try:
        wb = xlrd.open_workbook(filepath, on_demand=True)
        sheets_data = {}
        for sheet_name in wb.sheet_names():
            ws = wb.sheet_by_name(sheet_name)
            rows = []
            for row_idx in range(ws.nrows):
                row = []
                for col_idx in range(ws.ncols):
                    cell = ws.cell(row_idx, col_idx)
                    if cell.ctype == xlrd.XL_CELL_EMPTY:
                        val = None
                    elif cell.ctype == xlrd.XL_CELL_TEXT:
                        val = cell.value
                    elif cell.ctype == xlrd.XL_CELL_NUMBER:
                        val = cell.value
                    elif cell.ctype == xlrd.XL_CELL_DATE:
                        val = xlrd.xldate_as_datetime(cell.value, wb.datemode)
                    elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                        val = bool(cell.value)
                    else:
                        val = str(cell.value)
                    row.append(val)
                rows.append(row)
            sheets_data[sheet_name] = rows
        wb.release_resources()
        return sheets_data
    except Exception as e:
        raise RuntimeError(f"Failed to read {filepath}: {e}")


def read_csv(filepath: str) -> Dict[str, List[List[Any]]]:
    """Read CSV file, treating as single sheet"""
    try:
        rows = []
        # Try multiple encodings common for Excel CSV exports
        for encoding in ['utf-8-sig', 'utf-8', 'latin1', 'cp1252']:
            try:
                with codecs.open(filepath, 'r', encoding) as f:
                    reader = csv.reader(f)
                    rows = [row for row in reader]
                break
            except UnicodeDecodeError:
                continue
        if not rows:
            raise ValueError("Could not decode CSV with any supported encoding")
        return {"Sheet": rows}
    except Exception as e:
        raise RuntimeError(f"Failed to read CSV {filepath}: {e}")


def parse_excel_file(filepath: str) -> Dict[str, List[List[Any]]]:
    """Parse Excel or CSV file into dictionary of sheet_name -> rows"""
    ext = Path(filepath).suffix.lower()

    if ext == '.csv':
        return read_csv(filepath)
    elif ext == '.xls':
        return read_excel_xls(filepath)
    else:  # .xlsx, .xlsm, .xlsb
        return read_excel_xlsx(filepath)


def get_file_info(filepath: str) -> Dict[str, Any]:
    """Get file metadata"""
    p = Path(filepath)
    try:
        stat = p.stat()
        return {
            'name': p.name,
            'path': str(p.absolute()),
            'relative_path': str(p),
            'size': stat.st_size,
            'modified': stat.st_mtime,
            'modified_ms': stat_mtime_ms(stat)
        }
    except Exception:
        return {
            'name': p.name,
            'path': str(p.absolute()),
            'relative_path': str(p),
            'size': 0,
            'modified': 0
        }


def sheet_to_rows(sheet_data: List[List[Any]]) -> List[List[str]]:
    """Convert raw sheet data to cleaned rows (strings, None -> '')"""
    cleaned = []
    for row in sheet_data:
        cleaned_row = []
        for cell in row:
            if cell is None:
                cleaned_row.append("")
            else:
                cleaned_row.append(str(cell))
        cleaned.append(cleaned_row)
    return cleaned


def iter_excel_files(root_path: str) -> Iterator[tuple[str, Dict]]:
    """Iterate over all Excel files in directory, yielding (filepath, file_info)"""
    for fp in scan_directory(root_path):
        try:
            info = get_file_info(fp)
            yield fp, info
        except Exception as e:
            print(f"Error getting file info for {fp}: {e}", file=sys.stderr)
